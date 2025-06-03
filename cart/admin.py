from django.contrib import admin
from .models import Order, OrderItem

class OrderItemInline(admin.TabularInline):
    model = OrderItem
    extra = 0

@admin.register(Order)
class OrderAdmin(admin.ModelAdmin):
    list_display = ('id', 'name', 'user', 'created_at', 'status')
    list_filter = ('status', 'created_at')
    search_fields = ('name', 'phone', 'address')
    inlines = [OrderItemInline]
    actions = ['export_as_excel', 'mark_as_shipped']

    def export_as_excel(self, request, queryset):
        import openpyxl
        from django.http import HttpResponse
        from openpyxl.utils import get_column_letter
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Đơn hàng"

        headers = ['ID', 'Tên khách', 'SĐT', 'Địa chỉ', 'Trạng thái', 'Ngày tạo']
        for col_num, column_title in enumerate(headers, 1):
            ws[f"{get_column_letter(col_num)}1"] = column_title

        for row_num, order in enumerate(queryset, 2):
            ws[f"A{row_num}"] = order.id
            ws[f"B{row_num}"] = order.name
            ws[f"C{row_num}"] = order.phone
            ws[f"D{row_num}"] = order.address
            ws[f"E{row_num}"] = order.get_status_display()
            ws[f"F{row_num}"] = order.created_at.strftime('%Y-%m-%d %H:%M')

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        response = HttpResponse(stream.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=orders.xlsx'
        return response

    export_as_excel.short_description = "📥 Xuất đơn hàng ra Excel"

    def mark_as_shipped(self, request, queryset):
        updated = queryset.update(status='shipped')
        self.message_user(request, f"{updated} đơn hàng đã được đánh dấu là 'Đã giao'.")
    
    mark_as_shipped.short_description = "📦 Đánh dấu là đã giao"
